"""
Excel Formula Recalculation Script
Recalculates all formulas in an Excel file using the pure-Python `formulas` library.
Falls back to static analysis when the formulas library cannot handle certain functions.
"""

import json
import math
import os
import re
import signal
import sys
import subprocess
import xml.etree.ElementTree as ET
import zipfile
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation as _DV

# WPS/Kdocs exports add a non-standard 'id' attribute to <dataValidation> nodes.
# openpyxl strictly rejects unknown kwargs, so we patch to tolerate them.
_dv_orig_init = _DV.__init__

def _dv_patched_init(self, *args, **kwargs):
    kwargs.pop('id', None)
    _dv_orig_init(self, *args, **kwargs)

_DV.__init__ = _dv_patched_init

EXCEL_ERRORS = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']


# WPS-proprietary pseudo-formulas that are not real calculations (e.g. embedded images).
_WPS_PSEUDO_FORMULAS = ('=DISPIMG(', '=_XLFN.DISPIMG(')


def _decode_process_output(data) -> str:
    """Decode subprocess pipe output robustly across UTF-8/GBK environments."""
    if data is None:
        return ''
    if isinstance(data, str):
        return data
    for enc in ('utf-8', 'gb18030'):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode('utf-8', errors='replace')


def _count_formulas(filename: str) -> tuple[int, list[tuple[str, str, str]]]:
    """Return (count, [(sheet, coordinate, formula_text), ...])."""
    wb = load_workbook(filename, data_only=False)
    cells = []
    for sheet_name in wb.sheetnames:
        for row in wb[sheet_name].iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    upper = cell.value.upper()
                    if any(upper.startswith(p) for p in _WPS_PSEUDO_FORMULAS):
                        continue
                    cells.append((sheet_name, cell.coordinate, cell.value))
    wb.close()
    return len(cells), cells


def _parse_solution_key(key: str) -> str | None:
    """Extract 'Sheet!Cell' location from a formulas solution key.

    Keys look like: "'[file.xlsx]Sheet1'!B2" or "'[file.xlsx]Sheet1'!A1:A3"
    Returns "Sheet1!B2" or None if the key is a range (not a single cell).
    """
    m = re.search(r"\](.+?)'!([A-Z]+\d+)$", key)
    if m:
        return f'{m.group(1)}!{m.group(2)}'
    return None


def _register_namespaces(xml_bytes: bytes) -> None:
    """Register all XML namespace prefixes found in the document."""
    for event, data in ET.iterparse(BytesIO(xml_bytes), events=['start-ns']):
        prefix, uri = data
        try:
            ET.register_namespace(prefix, uri)
        except ValueError:
            pass


def _inject_cached_values(filepath: str, cache_map: dict[tuple[str, str], object]) -> None:
    """Write calculated values into xlsx formula cells as cached values.

    Modifies the xlsx file in-place by adding <v> elements alongside <f> elements
    in the sheet XML, so programs reading with data_only=True see the values.
    """
    if not cache_map:
        return

    NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    RELS_NS = 'http://schemas.openxmlformats.org/package/2006/relationships'
    RELS_OFFICEREL = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

    with zipfile.ZipFile(filepath, 'r') as zin:
        file_list = zin.namelist()
        contents = {name: zin.read(name) for name in file_list}

    wb_path = 'xl/workbook.xml'
    rels_path = 'xl/_rels/workbook.xml.rels'
    if wb_path not in contents or rels_path not in contents:
        return

    _register_namespaces(contents[wb_path])
    wb_root = ET.fromstring(contents[wb_path])
    sheets_el = wb_root.find(f'{{{NS}}}sheets')
    if sheets_el is None:
        return
    sheet_rid = {}
    for s in sheets_el.findall(f'{{{NS}}}sheet'):
        name = s.get('name')
        rid = s.get(f'{{{RELS_NS}}}id') or s.get(f'{{{RELS_OFFICEREL}}}id')
        if name and rid:
            sheet_rid[name] = rid

    _register_namespaces(contents[rels_path])
    rels_root = ET.fromstring(contents[rels_path])
    rid_file = {}
    for rel in rels_root:
        rid = rel.get('Id')
        target = rel.get('Target')
        if rid and target:
            rid_file[rid] = f'xl/{target}' if not target.startswith('/') else target.lstrip('/')

    by_sheet: dict[str, dict[str, object]] = {}
    for (sheet_name, cell_ref), value in cache_map.items():
        by_sheet.setdefault(sheet_name, {})[cell_ref] = value

    name_upper_to_actual = {name.upper(): name for name in sheet_rid}

    modified_any = False
    for sheet_name, cell_values in by_sheet.items():
        actual_name = name_upper_to_actual.get(sheet_name.upper())
        if not actual_name:
            continue
        rid = sheet_rid[actual_name]
        xml_path = rid_file.get(rid)
        if not xml_path or xml_path not in contents:
            continue

        raw = contents[xml_path]
        _register_namespaces(raw)
        root = ET.fromstring(raw)
        sheet_data = root.find(f'{{{NS}}}sheetData')
        if sheet_data is None:
            continue

        modified = False
        for row_el in sheet_data.findall(f'{{{NS}}}row'):
            for cell_el in row_el.findall(f'{{{NS}}}c'):
                ref = cell_el.get('r')
                if ref not in cell_values:
                    continue
                f_el = cell_el.find(f'{{{NS}}}f')
                if f_el is None:
                    continue

                value = cell_values[ref]
                v_el = cell_el.find(f'{{{NS}}}v')
                if v_el is None:
                    v_el = ET.SubElement(cell_el, f'{{{NS}}}v')

                is_err = any(err in str(value) for err in EXCEL_ERRORS)
                if is_err:
                    cell_el.set('t', 'e')
                    v_el.text = str(value)
                elif isinstance(value, bool):
                    cell_el.set('t', 'b')
                    v_el.text = '1' if value else '0'
                elif isinstance(value, (int, float)):
                    if 't' in cell_el.attrib:
                        del cell_el.attrib['t']
                    v_el.text = str(float(value))
                elif isinstance(value, str):
                    cell_el.set('t', 'str')
                    v_el.text = value
                else:
                    v_el.text = str(value)

                modified = True

        if modified:
            buf = BytesIO()
            ET.ElementTree(root).write(buf, xml_declaration=True, encoding='UTF-8')
            contents[xml_path] = buf.getvalue()
            modified_any = True

    if modified_any:
        out = BytesIO()
        with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as zout:
            for name in file_list:
                zout.writestr(name, contents[name])
        tmp_path = f'{filepath}.recalc_tmp'
        try:
            with open(tmp_path, 'wb') as f:
                f.write(out.getvalue())
                f.flush()
                os.fsync(f.fileno())
            os.replace(tmp_path, filepath)
        except BaseException:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
            raise


def _calculate_and_check(abs_path: str) -> tuple[str | None, int, dict]:
    """Calculate formulas, inject cached values, and detect errors.

    Returns (calc_error_msg_or_None, total_errors, error_summary).
    If calc_error_msg is not None, the calculation failed entirely.
    """
    try:
        import formulas
    except ImportError:
        return 'formulas library not installed (pip install formulas)', 0, {}

    # Build set of cells containing WPS pseudo-formulas (e.g. DISPIMG) to skip.
    pseudo_cells: set[str] = set()
    try:
        wb_tmp = load_workbook(abs_path, data_only=False)
        for sn in wb_tmp.sheetnames:
            for row in wb_tmp[sn].iter_rows():
                for c in row:
                    if c.value and isinstance(c.value, str):
                        upper = c.value.upper()
                        if any(upper.startswith(p) for p in _WPS_PSEUDO_FORMULAS):
                            pseudo_cells.add(f'{sn}!{c.coordinate}')
        wb_tmp.close()
    except Exception:
        pass

    try:
        os.environ['TQDM_DISABLE'] = '1'
        xl_model = formulas.ExcelModel().loads(abs_path).finish()
        solution = xl_model.calculate()
    except Exception as e:
        return str(e), 0, {}

    import numpy as np
    error_details: dict[str, list[str]] = {err: [] for err in EXCEL_ERRORS}
    total_errors = 0
    cache_map: dict[tuple[str, str], object] = {}

    for key, ranges_obj in solution.items():
        location = _parse_solution_key(key)
        if location is None:
            continue
        if location in pseudo_cells:
            continue

        sheet_name, cell_ref = location.split('!', 1)

        try:
            values = np.asarray(ranges_obj.value).flat
        except Exception:
            continue

        for val in values:
            val_str = str(val)
            is_error = False
            for err in EXCEL_ERRORS:
                if err in val_str:
                    error_details[err].append(location)
                    total_errors += 1
                    is_error = True
                    cache_map[(sheet_name, cell_ref)] = val_str
                    break

            if not is_error and isinstance(val, float):
                if math.isinf(val):
                    error_details['#DIV/0!'].append(location)
                    total_errors += 1
                    cache_map[(sheet_name, cell_ref)] = '#DIV/0!'
                elif math.isnan(val):
                    error_details['#VALUE!'].append(location)
                    total_errors += 1
                    cache_map[(sheet_name, cell_ref)] = '#VALUE!'
                else:
                    cache_map[(sheet_name, cell_ref)] = val
            elif not is_error:
                try:
                    native = val.item() if hasattr(val, 'item') else val
                except Exception:
                    native = val
                if isinstance(native, (int, float, str, bool)):
                    cache_map[(sheet_name, cell_ref)] = native

    try:
        _inject_cached_values(abs_path, cache_map)
    except Exception:
        pass

    error_summary = {}
    for err_type, locations in error_details.items():
        if locations:
            error_summary[err_type] = {
                'count': len(locations),
                'locations': locations[:20],
            }
    return None, total_errors, error_summary


def _static_analyze(formula_cells: list[tuple[str, str, str]],
                    sheet_names: list[str]) -> tuple[int, dict]:
    """Lightweight structural checks when full calculation is unavailable.

    Checks:
    - Referenced sheet names exist in the workbook
    - Parentheses are balanced
    """
    error_details: dict[str, list[str]] = {}
    total_errors = 0

    sheet_ref_re = re.compile(r"(?:'([^']+)'|([A-Za-z_]\w*))!")

    for sheet_name, coord, formula in formula_cells:
        location = f'{sheet_name}!{coord}'

        if formula.count('(') != formula.count(')'):
            error_details.setdefault('syntax_error', []).append(location)
            total_errors += 1
            continue

        for m in sheet_ref_re.finditer(formula):
            ref_sheet = m.group(1) or m.group(2)
            if ref_sheet not in sheet_names:
                error_details.setdefault('#REF!', []).append(location)
                total_errors += 1
                break

    error_summary = {}
    for err_type, locations in error_details.items():
        if locations:
            error_summary[err_type] = {
                'count': len(locations),
                'locations': locations[:20],
            }
    return total_errors, error_summary


def _recalc_impl(filename: str) -> dict:
    """Core recalc logic executed inside the subprocess."""
    if not Path(filename).exists():
        return {'error': f'File {filename} does not exist'}

    abs_path = str(Path(filename).absolute())

    try:
        formula_count, formula_cells = _count_formulas(abs_path)
    except Exception as e:
        return {'error': f'Failed to read workbook: {e}'}

    if formula_count == 0:
        return {
            'status': 'success',
            'total_errors': 0,
            'total_formulas': 0,
            'error_summary': {},
        }

    calc_error, total_errors, error_summary = _calculate_and_check(abs_path)

    if calc_error is not None:
        wb_tmp = load_workbook(abs_path, data_only=False)
        sheet_names = wb_tmp.sheetnames
        wb_tmp.close()

        total_errors, error_summary = _static_analyze(formula_cells, sheet_names)
        return {
            'status': 'success' if total_errors == 0 else 'errors_found',
            'total_errors': total_errors,
            'total_formulas': formula_count,
            'error_summary': error_summary,
            'warning': f'Full recalculation unavailable, static analysis only: {calc_error}',
        }

    return {
        'status': 'success' if total_errors == 0 else 'errors_found',
        'total_errors': total_errors,
        'total_formulas': formula_count,
        'error_summary': error_summary,
    }


def recalc(filename: str, timeout: int = 30) -> dict:
    """Recalculate all formulas and scan for errors.

    Runs the heavy calculation in a **separate Python process** via
    ``subprocess.Popen`` so that a hung ``formulas`` library can be killed
    after *timeout* seconds.  Falls back to static analysis on timeout.

    Args:
        filename: Absolute path to the xlsx file.
        timeout: Max seconds before the subprocess is killed.

    Returns:
        dict with keys: status, total_errors, total_formulas, error_summary.
        On failure: dict with 'error' key.
    """
    if not Path(filename).exists():
        return {'error': f'File {filename} does not exist'}

    abs_path = str(Path(filename).absolute())
    script = os.path.abspath(__file__)

    stdout = ''
    stderr = ''
    try:
        popen_kwargs = {
            'stdout': subprocess.PIPE,
            'stderr': subprocess.PIPE,
            'text': False,
        }
        if os.name == 'posix':
            popen_kwargs['preexec_fn'] = os.setsid
        elif os.name == 'nt':
            # Create a dedicated process group so we can terminate the process tree.
            popen_kwargs['creationflags'] = getattr(subprocess, 'CREATE_NEW_PROCESS_GROUP', 0)

        proc = subprocess.Popen([sys.executable, script, abs_path], **popen_kwargs)
        stdout_raw, stderr_raw = proc.communicate(timeout=timeout)
        stdout = _decode_process_output(stdout_raw)
        stderr = _decode_process_output(stderr_raw)

        if proc.returncode == 0 and (stdout or '').strip():
            return json.loads(stdout)
        return {'error': f'Subprocess failed (rc={proc.returncode}): {stderr[:500]}'}
    except subprocess.TimeoutExpired:
        if os.name == 'posix':
            try:
                os.killpg(proc.pid, signal.SIGKILL)
            except Exception:
                proc.kill()
        elif os.name == 'nt':
            try:
                # /T kills child processes; /F forces termination.
                subprocess.run(
                    ['taskkill', '/PID', str(proc.pid), '/T', '/F'],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                    timeout=5,
                )
            except Exception:
                proc.kill()
        else:
            proc.kill()

        try:
            stdout_raw, stderr_raw = proc.communicate(timeout=5)
            stdout = _decode_process_output(stdout_raw)
            stderr = _decode_process_output(stderr_raw)
        except subprocess.TimeoutExpired:
            # Extremely defensive: ensure process is gone, then do a short final drain.
            try:
                proc.kill()
            except Exception:
                pass
            try:
                stdout_raw, stderr_raw = proc.communicate(timeout=1)
                stdout = _decode_process_output(stdout_raw)
                stderr = _decode_process_output(stderr_raw)
            except Exception:
                stdout, stderr = '', ''
        except Exception:
            stdout, stderr = '', ''

        try:
            formula_count, formula_cells = _count_formulas(abs_path)
            wb_tmp = load_workbook(abs_path, data_only=False)
            sheet_names = wb_tmp.sheetnames
            wb_tmp.close()
            total_errors, error_summary = _static_analyze(formula_cells, sheet_names)
            return {
                'status': 'success' if total_errors == 0 else 'errors_found',
                'total_errors': total_errors,
                'total_formulas': formula_count,
                'error_summary': error_summary,
                'warning': f'Calculation timed out after {timeout}s, static analysis only',
            }
        except Exception as e:
            return {'error': f'Timed out after {timeout}s, static analysis also failed: {e}'}
    except json.JSONDecodeError:
        return {'error': f'Invalid output from subprocess: {stdout[:200]}'}
    except Exception as e:
        return {'error': f'Calculation failed: {e}'}


def main():
    if len(sys.argv) < 2:
        print('Usage: python recalc.py <excel_file>')
        sys.exit(1)
    filename = sys.argv[1]
    print(json.dumps(_recalc_impl(filename), ensure_ascii=False))


if __name__ == '__main__':
    main()